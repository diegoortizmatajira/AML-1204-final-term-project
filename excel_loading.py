from openpyxl import load_workbook


class ExcelInput:

    def __init__(self, row, category_1, category_2, category_3, category_4, category_5, question, answer, show_in_menu,
                 validation):
        self.row = row
        self.category_1 = category_1
        self.category_2 = category_2
        self.category_3 = category_3
        self.category_4 = category_4
        self.category_5 = category_5
        self.question = question
        self.answer = answer
        self.show_in_menu = show_in_menu
        self.validation = validation


class ExcelHeader:
    name: str = 'Sample chatbot'
    description: str = 'This is a sample chatbot'
    welcome_message: str = 'Hello, how can I help you?'
    selection_message: str = 'Please tell me which option fits your question?'
    selection_continuation_message: str = 'There are other options that may help you...'
    formal_offering_message: str = 'Can I help you with anything else?'
    final_message: str = 'Thank you for using this chat, hope to see you soon'
    feedback_message: str = 'Please provide feedback... Was the answer I gave you useful?'


def load_excel_file(filename: str) -> (list[ExcelInput], ExcelHeader):
    wb = load_workbook(filename=filename, data_only=True)
    ws = wb['Q&A']

    excel_rows_list = []

    for row in ws.iter_rows():
        row_number, category_1, category_2, category_3, category_4, \
            category_5, question, answer, show_in_menu, validation = tuple(map(lambda r: r.value, row))

        excel_input_element = ExcelInput(row_number, category_1, category_2, category_3, category_4, category_5,
                                         question,
                                         answer, show_in_menu, validation)
        excel_rows_list.append(excel_input_element)

    # Ignores the first row as it contains the Headers
    return excel_rows_list[1:], ExcelHeader()
